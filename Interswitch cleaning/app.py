from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
import sys
from uuid import uuid4

from flask import Flask, abort, render_template_string, request, send_file
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from werkzeug.utils import secure_filename


APP_TITLE = "Interswitch Date Sorter"
TARGET_HEADER = "Local_Date_Time"
MAX_HEADER_SCAN_ROWS = 25
ALLOWED_SUFFIXES = {".xlsx"}
BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
SORTED_DIR = BASE_DIR / "sorted"

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024


PAGE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{{ title }}</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --text: #171a1f;
      --muted: #667085;
      --line: #d9dee7;
      --accent: #0f766e;
      --accent-dark: #0b5f59;
      --success-bg: #ecfdf3;
      --success-line: #abefc6;
      --success-text: #067647;
      --danger: #b42318;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      background: var(--bg);
      color: var(--text);
    }
    main {
      min-height: 100vh;
      display: grid;
      place-items: center;
      padding: 32px 16px;
    }
    .shell {
      width: min(760px, 100%);
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 28px;
      box-shadow: 0 16px 40px rgba(15, 23, 42, 0.08);
    }
    h1 {
      margin: 0 0 8px;
      font-size: 28px;
      line-height: 1.2;
      letter-spacing: 0;
    }
    p {
      margin: 0 0 24px;
      color: var(--muted);
      line-height: 1.5;
    }
    form {
      display: grid;
      gap: 18px;
    }
    label {
      display: grid;
      gap: 8px;
      color: #344054;
      font-weight: 700;
    }
    input[type="file"], select {
      min-height: 44px;
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      padding: 10px 12px;
      color: var(--text);
      font: inherit;
      font-weight: 400;
    }
    .row {
      display: grid;
      grid-template-columns: 1fr 220px;
      gap: 16px;
      align-items: end;
    }
    button {
      min-height: 46px;
      border: 0;
      border-radius: 6px;
      background: var(--accent);
      color: #fff;
      font: inherit;
      font-weight: 700;
      cursor: pointer;
    }
    button:hover { background: var(--accent-dark); }
    .alert, .success {
      margin-bottom: 18px;
      border-radius: 6px;
      padding: 12px 14px;
      line-height: 1.4;
    }
    .alert {
      border: 1px solid #fecdca;
      background: #fff4f2;
      color: var(--danger);
    }
    .success {
      border: 1px solid var(--success-line);
      background: var(--success-bg);
      color: var(--success-text);
    }
    .success a, .file-list a {
      color: var(--accent-dark);
      font-weight: 700;
    }
    .file-list {
      margin-top: 24px;
      border-top: 1px solid var(--line);
      padding-top: 18px;
    }
    .file-list h2 {
      margin: 0 0 12px;
      font-size: 18px;
      letter-spacing: 0;
    }
    .file-list ul {
      display: grid;
      gap: 10px;
      list-style: none;
      margin: 0;
      padding: 0;
    }
    .file-list li {
      display: flex;
      justify-content: space-between;
      gap: 12px;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 10px 12px;
      color: var(--muted);
    }
    .note {
      margin-top: 18px;
      font-size: 14px;
      color: var(--muted);
    }
    @media (max-width: 640px) {
      .shell { padding: 22px; }
      .row { grid-template-columns: 1fr; }
      h1 { font-size: 24px; }
    }
  </style>
</head>
<body>
  <main>
    <section class="shell">
      <h1>{{ title }}</h1>
      <p>Upload an .xlsx workbook and download a copy sorted by the Local_Date_Time column.</p>
      {% if error %}
        <div class="alert">{{ error }}</div>
      {% endif %}
      {% if success %}
        <div class="success">
          Sorted file saved locally. <a href="{{ success.url }}">Download {{ success.name }}</a>
        </div>
      {% endif %}
      <form method="post" enctype="multipart/form-data" id="sort-form">
        <label>
          Excel workbook
          <input type="file" name="workbook" accept=".xlsx" required>
        </label>
        <div class="row">
          <label>
            Sort order
            <select name="sort_order">
              <option value="asc">Oldest first</option>
              <option value="desc">Newest first</option>
            </select>
          </label>
          <button type="submit" id="submit-button">Sort and download</button>
        </div>
      </form>
      <div class="note">
        Files are stored in local folders named uploads and sorted. Supported date examples: 2026-05-22 22:37:58 and 5/24/2026 9:08.
      </div>
      {% if files %}
        <div class="file-list">
          <h2>Recent sorted files</h2>
          <ul>
            {% for file in files %}
              <li>
                <span>{{ file.name }}</span>
                <a href="{{ file.url }}">Download</a>
              </li>
            {% endfor %}
          </ul>
        </div>
      {% endif %}
    </section>
  </main>
  <script>
    const form = document.getElementById("sort-form");
    const button = document.getElementById("submit-button");

    form.addEventListener("submit", () => {
      button.textContent = "Sorting...";
      button.disabled = true;
      button.style.opacity = "0.75";
      button.style.cursor = "wait";
    });

    window.addEventListener("pageshow", () => {
      button.textContent = "Sort and download";
      button.disabled = false;
      button.style.opacity = "1";
      button.style.cursor = "pointer";
    });
  </script>
</body>
</html>
"""


DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y",
)


def normalize_header(value):
    return str(value or "").strip().casefold()


def parse_local_date_time(value, workbook):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, time.min)

    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value, workbook.epoch)
        except (TypeError, ValueError):
            return None
        return parsed if isinstance(parsed, datetime) else datetime.combine(parsed, time.min)

    text = str(value).strip()
    if not text:
        return None

    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            pass

    return None


def find_header_cell(worksheet):
    expected = normalize_header(TARGET_HEADER)
    max_row = min(worksheet.max_row, MAX_HEADER_SCAN_ROWS)

    for row in worksheet.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            if normalize_header(cell.value) == expected:
                return cell.row, cell.column

    raise ValueError(f"Could not find a '{TARGET_HEADER}' header in the first {max_row} rows.")


def row_is_empty(values):
    return all(value is None or str(value).strip() == "" for value in values)


def ensure_storage_dirs():
    UPLOAD_DIR.mkdir(exist_ok=True)
    SORTED_DIR.mkdir(exist_ok=True)


def recent_sorted_files(limit=10):
    ensure_storage_dirs()
    files = sorted(SORTED_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    return [
        {"name": path.name, "url": f"/download/{path.name}"}
        for path in files[:limit]
    ]


def unique_workbook_name(filename, suffix=""):
    safe_name = secure_filename(filename) or "workbook.xlsx"
    stem = Path(safe_name).stem
    extension = Path(safe_name).suffix.casefold() or ".xlsx"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    token = uuid4().hex[:8]
    return f"{stem}{suffix}_{timestamp}_{token}{extension}"


def sort_workbook(input_path, descending=False):
    print(f"[sort] received {input_path.name}", flush=True)
    workbook = load_workbook(input_path)

    worksheet = workbook.active
    header_row, date_column = find_header_cell(worksheet)
    print(f"[sort] using sheet={worksheet.title!r} header_row={header_row} date_column={date_column}", flush=True)

    valid_rows = []
    unreadable_rows = []

    data_rows = worksheet.iter_rows(
        min_row=header_row + 1,
        max_row=worksheet.max_row,
        max_col=worksheet.max_column,
        values_only=True,
    )

    for row_number, row_values in enumerate(data_rows, start=header_row + 1):
        if row_is_empty(row_values):
            continue

        sort_value = parse_local_date_time(row_values[date_column - 1], workbook)
        if sort_value is None:
            unreadable_rows.append((row_number, row_values))
        else:
            valid_rows.append((sort_value, row_number, row_values))

    valid_rows.sort(key=lambda item: (item[0], item[1]), reverse=descending)
    sorted_rows = [row_values for _, _, row_values in valid_rows]
    sorted_rows.extend(row_values for _, row_values in unreadable_rows)
    print(
        f"[sort] rows={len(sorted_rows)} valid_dates={len(valid_rows)} unreadable_dates={len(unreadable_rows)}",
        flush=True,
    )

    for offset, row_values in enumerate(sorted_rows, start=header_row + 1):
        for column, value in enumerate(row_values, start=1):
            worksheet.cell(row=offset, column=column, value=value)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    print("[sort] workbook ready", flush=True)
    return output


@app.get("/")
def index():
    return render_template_string(PAGE, title=APP_TITLE, error=None, success=None, files=recent_sorted_files())


@app.post("/")
def upload():
    uploaded_file = request.files.get("workbook")
    sort_order = request.form.get("sort_order", "asc")

    if not uploaded_file or not uploaded_file.filename:
        return render_template_string(
            PAGE,
            title=APP_TITLE,
            error="Please choose an .xlsx file.",
            success=None,
            files=recent_sorted_files(),
        ), 400

    suffix = Path(uploaded_file.filename).suffix.casefold()
    if suffix not in ALLOWED_SUFFIXES:
        return render_template_string(
            PAGE,
            title=APP_TITLE,
            error="Please upload an .xlsx file.",
            success=None,
            files=recent_sorted_files(),
        ), 400

    ensure_storage_dirs()
    upload_name = unique_workbook_name(uploaded_file.filename)
    uploaded_path = UPLOAD_DIR / upload_name
    uploaded_file.save(uploaded_path)

    try:
        output = sort_workbook(uploaded_path, descending=sort_order == "desc")
    except Exception as exc:
        return render_template_string(
            PAGE,
            title=APP_TITLE,
            error=str(exc),
            success=None,
            files=recent_sorted_files(),
        ), 400

    original_name = Path(uploaded_file.filename).stem
    download_name = unique_workbook_name(f"{original_name}.xlsx", suffix="_sorted")
    sorted_path = SORTED_DIR / download_name
    sorted_path.write_bytes(output.getvalue())

    return render_template_string(
        PAGE,
        title=APP_TITLE,
        error=None,
        success={"name": download_name, "url": f"/download/{download_name}"},
        files=recent_sorted_files(),
    )


@app.get("/download/<path:filename>")
def download(filename):
    safe_name = secure_filename(filename)
    if safe_name != filename:
        abort(404)

    file_path = SORTED_DIR / safe_name
    if not file_path.exists() or file_path.suffix.casefold() not in ALLOWED_SUFFIXES:
        abort(404)

    return send_file(
        file_path,
        as_attachment=True,
        download_name=file_path.name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    ensure_storage_dirs()
    if sys.stdout is None:
        sys.stdout = open("server.log", "a", encoding="utf-8")
    if sys.stderr is None:
        sys.stderr = open("server.err.log", "a", encoding="utf-8")
    app.run(host="127.0.0.1", port=5000, debug=False)
